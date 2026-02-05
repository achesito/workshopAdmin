from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.generics import get_object_or_404
from rest_framework.decorators import action
from django.db.models import Count, Sum
from django.utils.timezone import now
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from ...services import maintenanceService
from ...services import spareDetailsService
from ...services import CustomerVehicleService
from ...serializer import maintenanceSerializer
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from django.template.loader import get_template
from xhtml2pdf import pisa
import requests
import base64
from tallerapp.local import *

class MaintenanceViewSet(viewsets.ModelViewSet):
    queryset = maintenanceService.objects()
    serializer_class = maintenanceSerializer
    
    def list(self, request, *args, **kwargs):
        start = self.request.GET.get('start')
        order = self.request.GET.get('order')
        level = 0
        if (int(start) == 0):
            level = (maintenanceService.objects().count() % int(order)) if maintenanceService.objects().count() < int(order) else int(order)
        else:
            result = maintenanceService.objects().count() - (int(order) * int(start))
            level = int(order) + (int(order) * int(start)) if result > int(order) else maintenanceService.objects().count()
        items = maintenanceService.objects()[(int(order) * int(start)) : level]
        return Response({
            'items' : maintenanceSerializer(items, many=True).data,
            'total' : maintenanceService.objects().count()  
        })
    
    def create(self, request, *args, **kwargs):
        return super().create(request, *args, **kwargs)
    
    def update(self, request, *args, **kwargs):
        return super().update(request, *args, **kwargs)
   
    def destroy(self, request, *args, **kwargs):
        return super().destroy(request, *args, **kwargs)
    
    @action(detail=False, methods=['get'], url_path='mantVehicle')
    def getByPlaca(self, request):
        plate = request.GET.get('plate')
        start = request.GET.get('start')
        order = request.GET.get('order')
        level = 0
        if (int(start) == 0):
            level = (len(maintenanceService.getByplate(plate)) % int(order)) if len(maintenanceService.getByplate(plate)) < int(order) else int(order)
        else:
            result = len(maintenanceService.getByplate(plate)) - (int(order) * int(start))
            level = int(order) + (int(order) * int(start)) if result > int(order) else len(maintenanceService.getByplate(plate))
        maintenance = maintenanceService.getByplate(plate)[(int(order) * int(start)) : level]
        serializer = maintenanceSerializer(maintenance, many=True, context={'request' : request})
        return Response({
            'items' : serializer.data,
            'total' : len(maintenanceService.getByplate(plate))
        })
        
    def get_summary_reports(self, period, year, month=None):
        if period == 'monthly' and month:
            qs = self.queryset.filter(date__year=year, date__month=month)
        else:
            qs = self.queryset.filter(date__year=year)

        if not qs.exists():
            return None, qs

        reports = [{
            'customer': CustomerVehicleService.objects().filter(vehicle=m.vehicle).first().customer.name,
            'type': m.type,
            'task_status': m.task_status,
            'date': m.date.strftime('%Y-%m-%d'),
            'cost_job': float(m.cost_job),
            'total': float(m.total),
        } for m in qs.select_related('vehicle')]
        

        summary_status = list(
            qs.values('task_status')
              .annotate(quantity=Count('idMaintenance'))
              .order_by('-quantity')
        )

        vehicle_top = qs.values('vehicle__model') \
                         .annotate(quantity=Count('idMaintenance')) \
                         .order_by('-quantity') \
                         .first()

        maintenance_top = qs.values('type') \
                              .annotate(quantity=Count('idMaintenance')) \
                              .order_by('-quantity') \
                              .first()

        details = spareDetailsService.getByparam(qs)
        spare_top = details.values('spare__name') \
                               .annotate(total_uses=Sum('quantity')) \
                               .order_by('-total_uses') \
                               .first()

        date_top = qs.values('date') \
                      .annotate(quantity=Count('idMaintenance')) \
                      .order_by('-quantity') \
                      .first()

        return {
            'reports': reports,
            'summary_status': summary_status,
            'vehicle_top': vehicle_top,
            'maintenance_top': maintenance_top,
            'spare_top': spare_top,
            'date_top': date_top
        }, qs

    @action(detail=False, methods=['get'], url_path='report')
    def report(self, request):
        period = request.query_params.get('period', 'monthly')
        try:
            year = int(request.query_params.get('year'))
            month = int(request.query_params.get('month')) if period == 'monthly' or period == 'mensual' else None
        except (TypeError, ValueError):
            return Response({'error': 'invalid params'}, status=status.HTTP_400_BAD_REQUEST)

        data, qs = self.get_summary_reports(period, year, month)
        if data is None:
            return Response({'mesagge': 'No registration over this time lapse'}, status=status.HTTP_404_NOT_FOUND)
        
        return Response({
            'period': period,
            'year': year,
            'month': month,
            **data,
            'version': now().timestamp(),
        })
    
    @action(detail=False, methods=['get'], url_path='Export_excel')
    def export_excel(self, request):
        period = request.query_params.get('period', 'monthly')
        try:
            year = int(request.query_params.get('year'))
            month = int(request.query_params.get('month')) if period == 'monthly' else None
        except (TypeError, ValueError):
            return Response({'error': 'invalid params'}, status=status.HTTP_400_BAD_REQUEST)

        data, qs = self.get_summary_reports(period, year, month)
        if data is None:
            return Response({'mesagge': 'No registration over this time lapse'}, status=status.HTTP_404_NOT_FOUND)

        langue = request.query_params.get('language')
        # Crear libro y hoja 
        wb = Workbook()
        ws = wb.active
        ws.title = "Maintenance Report" if str(langue) == "en" else "Reporte de mantenimientos"

        # Estilos
        title_style = Font(bold=True, size=14)
        header_style = Font(bold=True, color="FFFFFF")
        header_fill = "4F81BD"
        align_center = Alignment(horizontal="center")

        # Título del Report
        title = f"Report {'monthly' if period == 'monthly' else 'Annual'} - {year}" + (f" / {month:02d}" if month else "")
        ws.merge_cells('A1:F1')
        ws['A1'] = title
        ws['A1'].font = title_style
        ws['A1'].alignment = align_center

        row = 3  # Desde la fila 3

        # Insertar Resumen antes de la tabla
        ws[f'A{row}'] = "summary status" if str(langue) == "en" else "resumen de estados"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        header_summary = []
        if str(langue) == "en":
            header_summary = ["status", "quantity"]
        else:
            header_summary = ["estado", "cantidad"]
        
        ws.append(header_summary)
        for summary in data['summary_status']:
            ws.append([summary['task_status'], summary['quantity']])
            row += 1

        row += 1
        ws[f'A{row}'] = "Most common car" if str(langue) == "en" else "vehiculo mas común"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = data['vehicle_top']['vehicle__model'] if data['vehicle_top'] else "N/A"
        row += 1

        ws[f'A{row}'] = "Most common maintenance" if str(langue) == "en" else "mantenimiento mas solicitado"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = data['maintenance_top']['type'] if data['maintenance_top'] else "N/A"
        row += 1

        ws[f'A{row}'] = "Most used part" if str(langue) == "en" else "repuesto mas utilizado"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = data['spare_top']['spare__name'] if data['spare_top'] else "N/A"
        row += 1

        ws[f'A{row}'] = "Date with more arrivals" if str(langue) == "en" else "fecha con mas llegadas"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = data['date_top']['date'].strftime('%Y-%m-%d') if data['date_top'] else "N/A"
        row += 2
        
        meta = []
        if str(langue) == "en":
            meta = ["Customer", "Type Maintenance", "Task status", "Date", "Cost", "Total"]
        else:
            meta = ["Cliente", "tipo de mantenimiento", "estado", "fecha", "costo de servicio", "total"]
        # Encabezado de tabla
        ws.append(meta)
        header_row = ws[row]
        for cell in header_row:
            cell.font = header_style
            cell.fill = openpyxl.styles.PatternFill("solid", fgColor=header_fill)
            cell.alignment = align_center
        row += 1

        # Insertar data
        for r in data['reports']:
            ws.append([
                r['customer'],
                r['type'],
                r['task_status'],
                r['date'],
                r['cost_job'],
                r['total']
            ])

        # Ajustar anchos de columna automáticamente
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

        # Guardar respuesta
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"Report_{period}_{year}" + (f"_{month:02d}" if month else "") + ".xlsx"
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        return response
    
    @action(detail=False, methods=['get'], url_path='agenda')
    def agenda(self, request):
        items = maintenanceService.getAgenda()
        start = request.GET.get('start')
        order = request.GET.get('order')
        level = 0
        if (int(start) == 0):
            level = (len(items) % int(order)) if len(items) < int(order) else int(order)
        else:
            result = len(items) - (int(order) * int(start))
            level = int(order) + (int(order) * int(start)) if result > int(order) else len(items)
        elements = items[(int(order) * int(start)) : level]
        return Response({
            'items' : elements,
            'total' : len(items)
        })
    
    @action(detail=False,methods=['get'], url_path='future')
    def future(self, request):
        return Response(maintenanceService.futureMaints())    
    
    def createBill(self,request):
        temp = get_template('billPdf.html')
        html = temp.render({'user':request.user})
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="billPdf.pdf"'
        
        pisa.CreatePDF(
           html,
           dest=response,
           encoding='UTF-8' 
        )
        return response
    
    @action(detail=False,methods=['post'], url_path='mail')
    def send_mail(self, request):
        attachs = []
        to_email = request.POST.get('customer')
        subject = request.POST.get('affair')
        html_content = request.POST.get('content')
        sender = request.POST.get('sender')
        if (len(request.data) > 4):
            filepath = request.FILES.getlist("files")
            for file in filepath:
                attachs.append({
                    "name" : file.name,
                    #solo comprimido base64 el archivo puede ser subido
                    "content" : base64.b64encode(file.read()).decode('utf-8')
                })
        BREVO_URL = "https://api.brevo.com/v3/smtp/email"
        headers = {
            "api-key" : EMAIL_KEY,
            "Content-Type" : "multipart/form-data",
            "Accept" : "multipart/form-data",
            }
        payload = {
            "sender" : {
                "name" : "workshopAdmin",
                "email" : "aurisaoficial@gmail.com"
                },
            "to" : [
                {"email" : to_email}
                ],
            "replyTo": {
                "email" : str(sender),
                "name" : "workshop"
            },
            "subject" : subject,
            "htmlContent" : html_content,
        }
        if (len(request.data) > 4):
            payload.update({
                "attachment" : attachs
            })
        response = requests.post(
            BREVO_URL,
            json=payload,
            headers=headers,
            timeout=10
        )
        
        if response.status_code not in (200,201):
            raise Exception(
                f"Brevo error {response.status_code}: {response.text}"
        )   
        return Response(response)
    
    @action(detail=False, methods=['get'], url_path='filters')
    def getByFilters(self, request):
        start = request.GET.get('start')
        order = request.GET.get('order')
        total = request.GET.get('total')
        type = request.GET.get('type')
        dateFrom = request.GET.get('dateFrom')
        dateTo = request.GET.get('dateTo')
        model = request.GET.get('model')
        plate = request.GET.get('plate')
        customer = request.GET.get('customer')
        parameters = []
        query = """
        SELECT m.diagnosis, m.cost_job, m.task_status, m.date
        , m.type, m."idMaintenance", m.total, m.vehicle_id, m.payment_method
        FROM autotaller_maintenance m
        INNER JOIN autotaller_vehicle AS car 
        ON car.plate = m.vehicle_id
        INNER JOIN autotaller_customervehicle cv ON
        cv.vehicle_id = car.plate"""
        if (total != "any"):
            query += " WHERE total > %s AND total <= %s"
            parameters.append(str(total).split("-")[0])
            parameters.append(str(total).split("-")[1])
        if (type != "any"):
            query += " AND type = %s"
            parameters.append(type)
        if (dateTo != "" and dateTo != "NULL"):
            parameters.append(dateFrom)
        if (dateFrom != "" and dateFrom != "NULL"):
            parameters.append(dateTo)        
        #"OR date >= %s AND date <= %s"
        line = lambda x, y : "AND date >= %s" if str(x) != "" and str(y) == "" else " AND date <= %s" if x == "" and str(y) != "" else " AND date >= %s" + " AND date <= %s"  if str(x) != "" and str(y) != "" else "" 
        if (len(line(dateFrom, dateTo))):
            query += line(dateFrom, dateTo)
        if (model != "" and model != "/"):
            parameters.append(model)
            query += " AND car.model = %s"
        if (plate != "undefined" and plate != "" and plate != "/"):
            query += " AND car.plate = %s"
        if (customer != "" and customer != None):
            query += "AND cv.customer_id = %s"
            parameters.append(customer)
        elements = maintenanceService.objects().raw(query, parameters)
        serializer = maintenanceSerializer(elements,many=True, context={'request' : request})
        if (int(start) == 0):
            level = (len(elements) % int(order)) if len(elements) < int(order) else int(order)
        else:
            result = len(elements) - (int(order) * int(start))
            level = int(order) + (int(order) * int(start)) if result > int(order) else len(maintenanceService.objects().raw(query, parameters))
        data = maintenanceService.objects().raw(query, parameters)[(int(order) * int(start)) : level]
        serializer = self.get_serializer(data, many=True)
        return Response({
            'items' : serializer.data,
            'total' : len(maintenanceService.objects().raw(query, parameters))
        })